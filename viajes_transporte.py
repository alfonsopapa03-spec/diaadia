import streamlit as st
import psycopg2
import pandas as pd
from datetime import datetime, timedelta, time
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pytz

# ==================== CONFIGURACIÓN ====================
st.set_page_config(
    page_title="Control de Viajes",
    layout="wide",
    page_icon="🚚",
    initial_sidebar_state="collapsed"
)

# ==================== CREDENCIALES ====================
SUPABASE_DB_URL = "postgresql://postgres.hhzuggxvdzzfmnvfulmp:Negritasantia@aws-1-us-east-1.pooler.supabase.com:6543/postgres"

# ==================== CATÁLOGO PLACAS / CONDUCTORES ====================
PLACA_CONDUCTOR = {
    "NOX459": "HABID CAMACHO",
    "NOX460": "JOSE ORTEGA PEREZ",
    "NOX461": "CARLOS TAFUR",
    "SON047": "ISAIAS VESGA",
    "SON048": "FLAVIO ROSENDO MALTE TUTALCHA",
    "SOP148": "SLITH JOSE ORTEGA PACHECO",
    "SOP149": "ABRAHAM SEGUNDO ALVAREZ VALLE",
    "SOP150": None,
    "SRO661": "JULIAN CALETH CORONADO",
    "SRO672": "PEDRO VILLAMIL",
    "TMW882": "JESUS DAVID MONTES MOSQUERA",
    "TRL282": "CHRISTIAN MARTINEZ NAVARRO",
    "TRL298": "YEIMI DUQUE ZULUAGA",
    "UYQ308": "REIMUR MANUEL",
    "UYV084": "RAMON TAFUR HERNANDEZ",
    "UYY788": "EDUARDO RAFAEL OLIVARES ALCAZAR",
    "PSX350": "EDGAR DE JESUS RAMIREZ",
}

TODOS_CONDUCTORES = sorted([
    "REIMUR MANUEL",
    "HABID CAMACHO",
    "JOSE ORTEGA PEREZ",
    "CARLOS TAFUR",
    "ISAIAS VESGA",
    "FLAVIO ROSENDO MALTE TUTALCHA",
    "SLITH JOSE ORTEGA PACHECO",
    "ABRAHAM SEGUNDO ALVAREZ VALLE",
    "RAMON TAFUR HERNANDEZ",
    "JULIAN CALETH CORONADO",
    "PEDRO VILLAMIL",
    "JESUS DAVID MONTES MOSQUERA",
    "CHRISTIAN MARTINEZ NAVARRO",
    "YEIMI DUQUE ZULUAGA",
    "EDGAR DE JESUS RAMIREZ",
    "EDUARDO RAFAEL OLIVARES ALCAZAR",
])

ESTADOS_VIAJE = ["✅ Completado", "❌ Anulado", "⚠️ Incumplido", "🔄 En Curso"]

# ==================== RUTAS FRECUENTES ====================
RUTAS_FRECUENTES = [
    ("PUERTO PALERMO", "AGOFER"),
    ("PUERTO BARRANQUILLA", "VIA40"),
    ("PUERTO BARRANQUILLA", "PROCAR"),
    ("PUERTO BARRANQUILLA", "CIENAGA"),
    ("PUERTO BARRANQUILLA", "MEICO"),
    ("PUERTO BARRANQUILLA", "MEICO CIRCUNVALAR"),
    ("PUERTO BARRANQUILLA", "SOLEDAD"),
    ("PUERTO PALERMO", "ZF BAQ"),
    ("PUERTO BARRANQUILLA", "ZF BAQ"),
    ("ZF BAQ", "ZF BAQ"),
    ("ZF BAQ", "JUAN MINA"),
    ("ZF BAQ", "TRIANGULO"),
    ("PUERTO BARRANQUILLA", "JUAN MINA"),
    ("PUERTO BARRANQUILLA", "ALMAGRARIO"),
    ("PUERTO BARRANQUILLA", "ALPOPULAR"),
    ("PUERTO BARRANQUILLA", "AGOFER"),
    ("PUERTO BARRANQUILLA", "AGUACHICA"),
    ("PUERTO BARRANQUILLA", "IMPORTADO"),
    ("PUERTO BARRANQUILLA", "GALAPA"),
    ("PUERTO BARRANQUILLA", "CAYENAS"),
    ("PUERTO BARRANQUILLA", "OMEGA"),
    ("PUERTO BARRANQUILLA", "SANTA MARTA"),
    ("PUERTO BARRANQUILLA", "MEDELLIN"),
    ("PUERTO BARRANQUILLA", "MONTERIA"),
    ("PUERTO BARRANQUILLA", "PARAGUACHON"),
    ("PUERTO BARRANQUILLA", "SAN ROQUE"),
    ("PUERTO BARRANQUILLA", "VIA AEROPUERTO"),
    ("PUERTO BARRANQUILLA", "FRENTE AEROPUERTO"),
    ("PUERTO PALERMO", "CIRCUNVALAR"),
    ("PUERTO PALERMO", "MALAMBO"),
    ("PUERTO PALERMO", "MONTERIA"),
    ("CENTRO LOGISTICO CARTAGENA", "YARA"),
    ("CARTAGENA", "BARRANCABERMEJA"),
    ("PALMAR", "CARTAGENA"),
    ("MALAMBO", "MONTERIA"),
    ("PALERMO", "MALAMBO"),
]

ORIGENES_FRECUENTES = sorted(set(r[0] for r in RUTAS_FRECUENTES))
LABEL_MANUAL = "✏️ Escribir manualmente..."

# ==================== CLIENTES FRECUENTES ====================
CLIENTES_FRECUENTES = [
    "AGOFER",
    "MONOMEROS COLOMBO VENEZOLANOS S.A.",
    "PROCAR",
    "MEICO",
    "WORLD",
    "TRAIDING",
    "MAT2",
    "SULOGISTICS",
    "SUDECO",
    "TRIANGULO",
    "DELTA",
    "CARGO ANDINA",
    "TRANSOLICAR",
    "TLC",
    "TULUA MADERAS",
    "KBINA",
    "KABIBA",
    "PASIFIC",
    "MOTOTRANSPORTAMO",
]
LABEL_MANUAL_CLI = "✏️ Escribir manualmente..."

# ==================== COORDENADAS POR LUGAR ====================
COORDENADAS = {
    "PUERTO BARRANQUILLA":      (10.9831, -74.7894),
    "PUERTO PALERMO":           (10.9125, -74.7489),
    "PALERMO":                  (10.9125, -74.7489),
    "ZF BAQ":                   (10.9700, -74.8100),
    "AGOFER":                   (10.9190, -74.8010),
    "MEICO":                    (10.9650, -74.8350),
    "MEICO CIRCUNVALAR":        (10.9680, -74.8320),
    "PROCAR":                   (10.9550, -74.8200),
    "VIA40":                    (10.9900, -74.8000),
    "VIA AEROPUERTO":           (10.9990, -74.7780),
    "FRENTE AEROPUERTO":        (10.9990, -74.7780),
    "SOLEDAD":                  (10.9180, -74.7670),
    "MALAMBO":                  (10.8610, -74.7730),
    "GALAPA":                   (10.9060, -74.8880),
    "JUAN MINA":                (10.9750, -74.9200),
    "ALMAGRARIO":               (10.9620, -74.8150),
    "ALPOPULAR":                (10.9600, -74.8180),
    "CAYENAS":                  (10.9580, -74.8220),
    "OMEGA":                    (10.9570, -74.8230),
    "CIRCUNVALAR":              (10.9640, -74.8060),
    "TRIANGULO":                (10.9660, -74.8080),
    "IMPORTADO":                (10.9640, -74.8100),
    "CIENAGA":                  (11.0060, -74.2510),
    "SANTA MARTA":              (11.2408, -74.1990),
    "SAN ROQUE":                (8.5310,  -73.5730),
    "AGUACHICA":                (8.3097,  -73.6197),
    "PARAGUACHON":              (11.3320, -72.3820),
    "MONTERIA":                 (8.7575,  -75.8812),
    "MEDELLIN":                 (6.2442,  -75.5812),
    "BARRANCABERMEJA":          (7.0653,  -73.8547),
    "CARTAGENA":                (10.3910, -75.4794),
    "CENTRO LOGISTICO CARTAGENA": (10.4061, -75.5100),
    "PALMAR":                   (10.7800, -75.1100),
    "YARA":                     (10.3850, -75.4950),
}

# ==================== CSS ====================
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@400;600;700&family=Barlow:wght@300;400;500&display=swap');
    html, body, [class*="css"] { font-family: 'Barlow', sans-serif; }
    .main-header {
        background: linear-gradient(135deg, #0f2027, #203a43, #2c5364);
        padding: 1.5rem 2rem; border-radius: 12px; margin-bottom: 1.5rem;
    }
    .main-header h1 {
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 2rem; font-weight: 700; color: white; margin: 0; letter-spacing: 1px;
    }
    .main-header p { color: #a0c4d8; margin: 0; font-size: 0.9rem; }
    .kpi-box {
        background: white; border-radius: 10px; padding: 1rem 1.2rem;
        border-left: 5px solid #2c5364; box-shadow: 0 2px 8px rgba(0,0,0,0.07);
        margin-bottom: 0.5rem;
    }
    .kpi-box .kpi-val { font-size: 2rem; font-weight: 700; color: #0f2027; }
    .kpi-box .kpi-lbl { font-size: 0.8rem; color: #666; text-transform: uppercase; letter-spacing: 1px; }
    div[data-testid="stTabs"] button {
        font-family: 'Barlow Condensed', sans-serif;
        font-weight: 600; font-size: 1rem; letter-spacing: 0.5px;
    }
    .conductor-auto {
        background: #e8f5e9; border-left: 4px solid #2ecc71;
        padding: 0.5rem 1rem; border-radius: 6px; margin: 0.3rem 0;
        font-weight: 600; color: #1a5c2a;
    }
    .conductor-manual {
        background: #fff3e0; border-left: 4px solid #f39c12;
        padding: 0.5rem 1rem; border-radius: 6px; margin: 0.3rem 0;
        font-weight: 600; color: #7d4600;
    }
    /* Estilos para el bloque de días extra */
    .dias-extra-box {
        background: #fff8e1;
        border: 1px solid #f9a825;
        border-radius: 8px;
        padding: 0.6rem 1rem;
        margin: 0.4rem 0;
    }
    .dias-extra-box label {
        color: #e65100 !important;
        font-weight: 600 !important;
    }
</style>
""", unsafe_allow_html=True)

# ==================== BASE DE DATOS ====================
class DB:
    def __init__(self):
        self.url = SUPABASE_DB_URL
        self.init()

    def conn(self):
        return psycopg2.connect(self.url)

    def init(self):
        try:
            c = self.conn()
            cur = c.cursor()
            cur.execute("""
                CREATE TABLE IF NOT EXISTS viajes_transporte (
                    id SERIAL PRIMARY KEY,
                    fecha_registro TIMESTAMP DEFAULT (now() AT TIME ZONE 'America/Bogota'),
                    fecha DATE NOT NULL,
                    placa TEXT NOT NULL,
                    conductor TEXT,
                    cliente TEXT,
                    origen TEXT,
                    destino TEXT,
                    hora_cita_cargue TIME,
                    hora_salida_cargue TIME,
                    hora_llegada_descargue TIME,
                    hora_salida_descargue TIME,
                    contenedor TEXT,
                    carga TEXT,
                    numero_importacion_bl TEXT,
                    manifiesto TEXT,
                    observacion TEXT,
                    estado TEXT DEFAULT 'Completado',
                    dias_salida_cargue INTEGER DEFAULT 0,
                    dias_llegada_descargue INTEGER DEFAULT 0,
                    dias_salida_descargue INTEGER DEFAULT 0
                )
            """)
            # Agregar columnas si no existen (migraciones)
            cols_add = [
                "ALTER TABLE viajes_transporte ADD COLUMN IF NOT EXISTS cliente TEXT",
                "ALTER TABLE viajes_transporte ADD COLUMN IF NOT EXISTS contenedor TEXT",
                "ALTER TABLE viajes_transporte ADD COLUMN IF NOT EXISTS numero_importacion_bl TEXT",
                "ALTER TABLE viajes_transporte ADD COLUMN IF NOT EXISTS manifiesto TEXT",
                "ALTER TABLE viajes_transporte ADD COLUMN IF NOT EXISTS estado TEXT DEFAULT 'Completado'",
                "ALTER TABLE viajes_transporte ADD COLUMN IF NOT EXISTS dias_salida_cargue INTEGER DEFAULT 0",
                "ALTER TABLE viajes_transporte ADD COLUMN IF NOT EXISTS dias_llegada_descargue INTEGER DEFAULT 0",
                "ALTER TABLE viajes_transporte ADD COLUMN IF NOT EXISTS dias_salida_descargue INTEGER DEFAULT 0",
            ]
            for col in cols_add:
                try:
                    cur.execute(col); c.commit()
                except Exception:
                    try: c.rollback()
                    except: pass
            c.commit(); c.close()
        except Exception as e:
            st.error(f"Error DB init: {e}")

    def guardar_viaje(self, datos: dict) -> bool:
        try:
            c = self.conn(); cur = c.cursor()
            cur.execute("""
                INSERT INTO viajes_transporte
                (fecha, placa, conductor, cliente, origen, destino,
                 hora_cita_cargue, hora_salida_cargue, hora_llegada_descargue, hora_salida_descargue,
                 contenedor, carga, numero_importacion_bl, manifiesto, observacion, estado,
                 dias_salida_cargue, dias_llegada_descargue, dias_salida_descargue)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                datos["fecha"], datos["placa"], datos["conductor"], datos["cliente"],
                datos["origen"], datos["destino"],
                datos["hora_cita_cargue"], datos["hora_salida_cargue"],
                datos["hora_llegada_descargue"], datos["hora_salida_descargue"],
                datos["contenedor"], datos["carga"],
                datos["numero_importacion_bl"], datos["manifiesto"],
                datos["observacion"], datos["estado"],
                datos.get("dias_salida_cargue", 0),
                datos.get("dias_llegada_descargue", 0),
                datos.get("dias_salida_descargue", 0),
            ))
            c.commit(); c.close()
            return True
        except Exception as e:
            st.error(f"Error guardando: {e}"); return False

    def actualizar_viaje(self, viaje_id: int, datos: dict) -> bool:
        try:
            c = self.conn(); cur = c.cursor()
            cur.execute("""
                UPDATE viajes_transporte SET
                fecha=%s, placa=%s, conductor=%s, cliente=%s, origen=%s, destino=%s,
                hora_cita_cargue=%s, hora_salida_cargue=%s,
                hora_llegada_descargue=%s, hora_salida_descargue=%s,
                contenedor=%s, carga=%s, numero_importacion_bl=%s,
                manifiesto=%s, observacion=%s, estado=%s,
                dias_salida_cargue=%s, dias_llegada_descargue=%s, dias_salida_descargue=%s
                WHERE id=%s
            """, (
                datos["fecha"], datos["placa"], datos["conductor"], datos["cliente"],
                datos["origen"], datos["destino"],
                datos["hora_cita_cargue"], datos["hora_salida_cargue"],
                datos["hora_llegada_descargue"], datos["hora_salida_descargue"],
                datos["contenedor"], datos["carga"],
                datos["numero_importacion_bl"], datos["manifiesto"],
                datos["observacion"], datos["estado"],
                datos.get("dias_salida_cargue", 0),
                datos.get("dias_llegada_descargue", 0),
                datos.get("dias_salida_descargue", 0),
                viaje_id
            ))
            c.commit(); c.close(); return True
        except Exception as e:
            st.error(f"Error actualizando: {e}"); return False

    def eliminar_viaje(self, viaje_id: int) -> bool:
        try:
            c = self.conn(); cur = c.cursor()
            cur.execute("DELETE FROM viajes_transporte WHERE id=%s", (viaje_id,))
            c.commit(); c.close(); return True
        except Exception as e:
            st.error(f"Error eliminando: {e}"); return False

    def obtener_viajes(self, fecha_ini=None, fecha_fin=None, placa=None,
                       conductor=None, cliente=None, estado=None) -> pd.DataFrame:
        c = self.conn()
        q = """SELECT id, fecha, placa, conductor, cliente, origen, destino,
                      hora_cita_cargue, hora_salida_cargue,
                      hora_llegada_descargue, hora_salida_descargue,
                      contenedor, carga, numero_importacion_bl,
                      manifiesto, observacion, estado,
                      COALESCE(dias_salida_cargue, 0) as dias_salida_cargue,
                      COALESCE(dias_llegada_descargue, 0) as dias_llegada_descargue,
                      COALESCE(dias_salida_descargue, 0) as dias_salida_descargue
               FROM viajes_transporte WHERE 1=1"""
        params = []
        if fecha_ini: q += " AND fecha >= %s"; params.append(fecha_ini)
        if fecha_fin: q += " AND fecha <= %s"; params.append(fecha_fin)
        if placa and placa != "Todas": q += " AND placa = %s"; params.append(placa)
        if conductor: q += " AND conductor ILIKE %s"; params.append(f"%{conductor}%")
        if cliente: q += " AND cliente ILIKE %s"; params.append(f"%{cliente}%")
        if estado and estado != "Todos": q += " AND estado = %s"; params.append(estado)
        q += " ORDER BY fecha DESC, id DESC"
        try:
            df = pd.read_sql(q, c, params=params); return df
        except: return pd.DataFrame()
        finally: c.close()

    def placas_unicas(self):
        c = self.conn()
        try:
            df = pd.read_sql("SELECT DISTINCT placa FROM viajes_transporte ORDER BY placa", c)
            return df["placa"].tolist()
        except: return []
        finally: c.close()

    def stats_dashboard(self, fecha_ini, fecha_fin):
        c = self.conn()
        try:
            df = pd.read_sql("""
                SELECT fecha, placa, conductor, cliente, estado,
                       hora_cita_cargue, hora_salida_cargue,
                       hora_llegada_descargue, hora_salida_descargue,
                       COALESCE(dias_salida_cargue, 0) as dias_salida_cargue,
                       COALESCE(dias_llegada_descargue, 0) as dias_llegada_descargue,
                       COALESCE(dias_salida_descargue, 0) as dias_salida_descargue
                FROM viajes_transporte
                WHERE fecha >= %s AND fecha <= %s
                ORDER BY fecha
            """, c, params=[fecha_ini, fecha_fin])
            return df
        except: return pd.DataFrame()
        finally: c.close()


# ==================== HELPERS ====================
def hora_a_time(val):
    if val is None or (isinstance(val, float) and pd.isna(val)): return None
    if isinstance(val, time): return val
    try:
        s = str(val)[:5]; h, m = s.split(":"); return time(int(h), int(m))
    except: return None

def str_hora(val):
    t = hora_a_time(val)
    return t.strftime("%H:%M") if t else "—"

def calcular_duracion(h_ini, h_fin, dias_extra=0):
    """
    Calcula duración en minutos entre dos horas, con soporte para días extra.
    dias_extra: cuántos días adicionales pasaron entre h_ini y h_fin (0 = mismo día).
    """
    t1 = hora_a_time(h_ini)
    t2 = hora_a_time(h_fin)
    if not t1 or not t2: return None
    d1 = timedelta(hours=t1.hour, minutes=t1.minute)
    d2 = timedelta(hours=t2.hour, minutes=t2.minute)
    diff = d2 - d1 + timedelta(days=int(dias_extra or 0))
    # Si el resultado es negativo y no se especificaron días extra, asumimos +1 día
    if diff.total_seconds() < 0 and (dias_extra or 0) == 0:
        diff += timedelta(days=1)
    return int(diff.total_seconds() / 60)

def mins_a_str(mins):
    if mins is None: return "—"
    mins = int(mins)
    dias = mins // (60 * 24)
    resto = mins % (60 * 24)
    h, m = divmod(resto, 60)
    if dias > 0:
        return f"{dias}d {h}h {m:02d}m"
    return f"{h}h {m:02d}m"


# ==================== WIDGET HORAS CON DÍAS EXTRA ====================
def widget_horas(prefix, label_cita, label_sal_cargue, label_ll_desc, label_sal_desc,
                 val_cita=None, val_sal_cargue=None, val_ll_desc=None, val_sal_desc=None,
                 dias_sc=0, dias_ld=0, dias_sd=0):
    """
    Renderiza los 4 campos de hora + 3 spinners de 'días extra'.
    Retorna: (hora_cita, hora_sal_cargue, hora_ll_desc, hora_sal_desc,
              dias_salida_cargue, dias_llegada_descargue, dias_salida_descargue)
    """
    st.markdown("#### ⏱️ Tiempos de Operación")
    st.caption("💡 Si el vehículo tardó más de un día en pasar de una etapa a la siguiente, indica los **+días extra** debajo de cada hora.")

    h1, h2, h3, h4 = st.columns(4)
    with h1:
        hora_cita = st.time_input(label_cita, value=val_cita, step=300, key=f"{prefix}_hcc")
    with h2:
        hora_sc = st.time_input(label_sal_cargue, value=val_sal_cargue, step=300, key=f"{prefix}_hsc")
    with h3:
        hora_ld = st.time_input(label_ll_desc, value=val_ll_desc, step=300, key=f"{prefix}_hld")
    with h4:
        hora_sd = st.time_input(label_sal_desc, value=val_sal_desc, step=300, key=f"{prefix}_hsd")

    # Fila de días extra — debajo de cada hora salvo la primera
    d1_spacer, d2, d3, d4 = st.columns(4)
    with d1_spacer:
        st.markdown("")  # espaciador visual bajo "Cita Cargue" (no tiene días extra)
    with d2:
        d_sc = st.number_input(
            "➕ Días extra Salida Cargue",
            min_value=0, max_value=30, value=int(dias_sc or 0), step=1,
            key=f"{prefix}_dsc",
            help="¿Cuántos días tardó el vehículo en salir del cargue desde la cita?"
        )
    with d3:
        d_ld = st.number_input(
            "➕ Días extra Llegada Descargue",
            min_value=0, max_value=30, value=int(dias_ld or 0), step=1,
            key=f"{prefix}_dld",
            help="¿Cuántos días tardó en llegar al descargue desde que salió del cargue?"
        )
    with d4:
        d_sd = st.number_input(
            "➕ Días extra Salida Descargue",
            min_value=0, max_value=30, value=int(dias_sd or 0), step=1,
            key=f"{prefix}_dsd",
            help="¿Cuántos días tardó en salir del descargue desde que llegó?"
        )

    # Preview de duraciones en tiempo real
    if hora_cita or hora_sc or hora_ld or hora_sd:
        t_espera   = calcular_duracion(hora_cita, hora_sc, d_sc)
        t_transito = calcular_duracion(hora_sc,   hora_ld, d_ld)
        t_descargue= calcular_duracion(hora_ld,   hora_sd, d_sd)
        t_total    = None
        if t_espera is not None and t_transito is not None and t_descargue is not None:
            t_total = t_espera + t_transito + t_descargue

        cols_prev = st.columns(4)
        previews = [
            ("⏳ Espera Cargue",  t_espera),
            ("🚛 Tránsito",       t_transito),
            ("📦 Descargue",      t_descargue),
            ("🕐 Total Operación",t_total),
        ]
        for col, (lbl, val) in zip(cols_prev, previews):
            color = "#1abc9c" if val is not None else "#bdc3c7"
            col.markdown(
                f"<div style='text-align:center; padding:6px; background:#f0f4f8; border-radius:6px;"
                f"border-top:3px solid {color}; margin-top:4px;'>"
                f"<div style='font-size:0.7rem;color:#666;'>{lbl}</div>"
                f"<div style='font-size:1.1rem;font-weight:700;color:{color};'>{mins_a_str(val)}</div>"
                f"</div>",
                unsafe_allow_html=True
            )

    return hora_cita, hora_sc, hora_ld, hora_sd, d_sc, d_ld, d_sd


# ==================== EXCEL ====================
def generar_excel(df: pd.DataFrame, titulo: str = "Control de Viajes") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Viajes"

    ft_titulo  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ft_header  = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    ft_normal  = Font(name="Calibri", size=9)
    ft_total   = Font(name="Calibri", bold=True, size=10)
    ft_anulado = Font(name="Calibri", size=9, color="C0392B")
    ft_incump  = Font(name="Calibri", size=9, color="D35400")

    fill_titulo  = PatternFill("solid", start_color="0F2027")
    fill_header  = PatternFill("solid", start_color="203A43")
    fill_alt     = PatternFill("solid", start_color="EBF5FB")
    fill_total   = PatternFill("solid", start_color="D5DBDB")
    fill_anulado = PatternFill("solid", start_color="FADBD8")
    fill_incump  = PatternFill("solid", start_color="FDEBD0")

    borde  = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"),  bottom=Side(style="thin"))
    centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    izq    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    ws.merge_cells("A1:P1")
    now_col = datetime.now(pytz.timezone("America/Bogota"))
    ws["A1"] = f"🚚 {titulo}   |   Generado: {now_col.strftime('%d/%m/%Y %H:%M')} (COL)   |   Total: {len(df)} viajes"
    ws["A1"].font = ft_titulo
    ws["A1"].fill = fill_titulo
    ws["A1"].alignment = centro
    ws.row_dimensions[1].height = 30

    columnas = [
        ("fecha","FECHA",12), ("placa","PLACA",12), ("conductor","CONDUCTOR",26),
        ("cliente","CLIENTE",22), ("origen","ORIGEN",20), ("destino","DESTINO",20),
        ("hora_cita_cargue","H.CITA CARGUE",14), ("hora_salida_cargue","H.SALIDA CARGUE",14),
        ("hora_llegada_descargue","H.LLEGADA DESC.",14), ("hora_salida_descargue","H.SALIDA DESC.",14),
        ("contenedor","CONTENEDOR",18), ("carga","CARGA",14),
        ("numero_importacion_bl","IMP / BL",18), ("manifiesto","MANIFIESTO",12),
        ("observacion","OBSERVACIÓN",28), ("estado","ESTADO",14),
    ]
    cols_tiempos = ["ESPERA CARGUE","TRÁNSITO","DESCARGUE","TOTAL OPERACIÓN"]
    cols_coord   = ["LAT. ORIGEN","LON. ORIGEN","LAT. DESTINO","LON. DESTINO"]
    total_cols = len(columnas) + len(cols_tiempos) + len(cols_coord)

    ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")

    for idx, (key, nombre, ancho) in enumerate(columnas, start=1):
        cell = ws.cell(row=2, column=idx, value=nombre)
        cell.font = ft_header; cell.fill = fill_header
        cell.alignment = centro; cell.border = borde
        ws.column_dimensions[get_column_letter(idx)].width = ancho

    col_t_start = len(columnas) + 1
    for i, nombre in enumerate(cols_tiempos, start=col_t_start):
        cell = ws.cell(row=2, column=i, value=nombre)
        cell.font = ft_header
        cell.fill = PatternFill("solid", start_color="1A5276")
        cell.alignment = centro; cell.border = borde
        ws.column_dimensions[get_column_letter(i)].width = 16

    col_c_start = col_t_start + len(cols_tiempos)
    for i, nombre in enumerate(cols_coord, start=col_c_start):
        cell = ws.cell(row=2, column=i, value=nombre)
        cell.font = ft_header
        cell.fill = PatternFill("solid", start_color="1A5276")
        cell.alignment = centro; cell.border = borde
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.row_dimensions[2].height = 28

    for row_idx, (_, fila) in enumerate(df.iterrows(), start=3):
        estado_val = str(fila.get("estado", ""))
        es_an = "Anulado" in estado_val
        es_in = "Incumplido" in estado_val
        fill_f = fill_anulado if es_an else (fill_incump if es_in else (fill_alt if row_idx % 2 == 0 else None))

        for col_idx, (key, _, _) in enumerate(columnas, start=1):
            val = fila.get(key, "")
            if not isinstance(val, str) and pd.isna(val): val = ""
            if key.startswith("hora_") and val:
                try: val = str(val)[:5]
                except: val = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val != "" else "")
            cell.border = borde
            cell.alignment = centro if key in ("fecha","placa","estado") or key.startswith("hora_") else izq
            cell.font = ft_anulado if es_an else (ft_incump if es_in else ft_normal)
            if fill_f: cell.fill = fill_f

        # Calcular tiempos con días extra
        d_sc = int(fila.get("dias_salida_cargue",  0) or 0)
        d_ld = int(fila.get("dias_llegada_descargue", 0) or 0)
        d_sd = int(fila.get("dias_salida_descargue", 0) or 0)

        t_espera   = calcular_duracion(fila.get("hora_cita_cargue"),       fila.get("hora_salida_cargue"),     d_sc)
        t_transito = calcular_duracion(fila.get("hora_salida_cargue"),     fila.get("hora_llegada_descargue"), d_ld)
        t_descargue= calcular_duracion(fila.get("hora_llegada_descargue"), fila.get("hora_salida_descargue"),  d_sd)
        t_total    = (t_espera + t_transito + t_descargue) if (t_espera and t_transito and t_descargue) else None

        fill_t = PatternFill("solid", start_color="D6EAF8") if fill_f is None and row_idx % 2 == 0 else fill_f
        for ci, val in enumerate([mins_a_str(t_espera), mins_a_str(t_transito),
                                   mins_a_str(t_descargue), mins_a_str(t_total)],
                                  start=col_t_start):
            cell = ws.cell(row=row_idx, column=ci, value=val)
            cell.font = ft_normal; cell.border = borde; cell.alignment = centro
            if fill_t: cell.fill = fill_t

        origen_v  = str(fila.get("origen",  "") or "").strip().upper()
        destino_v = str(fila.get("destino", "") or "").strip().upper()
        lat_o, lon_o = COORDENADAS.get(origen_v,  (None, None))
        lat_d, lon_d = COORDENADAS.get(destino_v, (None, None))
        for ci, val in enumerate([lat_o, lon_o, lat_d, lon_d], start=col_c_start):
            cell = ws.cell(row=row_idx, column=ci, value=val if val is not None else "")
            cell.font = ft_normal; cell.border = borde; cell.alignment = centro
            if fill_t: cell.fill = fill_t
        ws.row_dimensions[row_idx].height = 18

    completados = len(df[df["estado"].str.contains("Completado", na=False)]) if "estado" in df.columns else 0
    anulados    = len(df[df["estado"].str.contains("Anulado",    na=False)]) if "estado" in df.columns else 0
    incumplidos = len(df[df["estado"].str.contains("Incumplido", na=False)]) if "estado" in df.columns else 0

    total_row = len(df) + 3
    try:
        ws.merge_cells(f"A{total_row}:{get_column_letter(len(columnas))}{total_row}")
    except Exception:
        pass
    ct = ws.cell(row=total_row, column=1,
                 value=f"TOTAL VIAJES: {len(df)}   |   ✅ {completados}  ❌ {anulados}  ⚠️ {incumplidos}")
    ct.font = ft_total; ct.fill = fill_total; ct.alignment = centro

    # ==================== HOJA RESUMEN ====================
    ws2 = wb.create_sheet("Resumen")

    def hdr(ws, fila, col1, col2, texto):
        c = ws.cell(fila, col1, texto)
        c.font = ft_header
        c.fill = PatternFill("solid", start_color="203A43")
        c.alignment = centro
        c.border = borde
        ws.row_dimensions[fila].height = 20
        for col in range(col1+1, col2+1):
            cx = ws.cell(fila, col, "")
            cx.fill = PatternFill("solid", start_color="203A43")
            cx.border = borde

    ws2["A1"] = "Resumen General de Operaciones"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", start_color="0F2027")
    ws2["A1"].alignment = centro
    ws2.row_dimensions[1].height = 26

    hdr(ws2, 2, 1, 2, "RESUMEN GENERAL")
    en_curso = len(df[df["estado"].str.contains("En Curso", na=False)]) if "estado" in df.columns else 0
    kpis = [
        ("Total Viajes", len(df)),
        ("Completados", completados),
        ("Anulados", anulados),
        ("Incumplidos", incumplidos),
        ("En Curso", en_curso),
        ("% Cumplimiento", f"{round(completados/len(df)*100,1)}%" if len(df) > 0 else "0%"),
    ]
    for i, (m, v) in enumerate(kpis, start=3):
        c1 = ws2.cell(i, 1, m); c2 = ws2.cell(i, 2, v)
        c1.font = ft_normal; c2.font = ft_total
        c1.border = borde; c2.border = borde
        c1.alignment = izq; c2.alignment = centro
        if i % 2 == 0:
            c1.fill = PatternFill("solid", start_color="EBF5FB")
            c2.fill = PatternFill("solid", start_color="EBF5FB")

    if "cliente" in df.columns and df["cliente"].notna().any():
        hdr(ws2, 2, 4, 5, "VIAJES POR CLIENTE")
        por_cli = df.groupby("cliente").size().reset_index(name="v").sort_values("v", ascending=False)
        for i, row in enumerate(por_cli.itertuples(), start=3):
            c1 = ws2.cell(i, 4, row.cliente); c2 = ws2.cell(i, 5, int(row.v))
            c1.font = ft_normal; c2.font = ft_total
            c1.border = borde; c2.border = borde
            c1.alignment = izq; c2.alignment = centro
            if i % 2 == 0:
                c1.fill = PatternFill("solid", start_color="EBF5FB")
                c2.fill = PatternFill("solid", start_color="EBF5FB")

    if "placa" in df.columns:
        hdr(ws2, 2, 7, 8, "VIAJES POR PLACA")
        por_placa = df.groupby("placa").size().reset_index(name="v").sort_values("v", ascending=False)
        for i, row in enumerate(por_placa.itertuples(), start=3):
            c1 = ws2.cell(i, 7, row.placa); c2 = ws2.cell(i, 8, int(row.v))
            c1.font = ft_normal; c2.font = ft_total
            c1.border = borde; c2.border = borde
            c1.alignment = izq; c2.alignment = centro
            if i % 2 == 0:
                c1.fill = PatternFill("solid", start_color="EBF5FB")
                c2.fill = PatternFill("solid", start_color="EBF5FB")

    for col_l, w in zip(["A","B","C","D","E","F","G","H"], [22,10,3,24,8,3,12,8]):
        ws2.column_dimensions[col_l].width = w

    # ==================== HOJA CONDUCTORES ====================
    ws3 = wb.create_sheet("Conductores")
    ws3["A1"] = "Ranking de Conductores"
    ws3["A1"].font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    ws3["A1"].fill = PatternFill("solid", start_color="0F2027")
    ws3["A1"].alignment = centro
    ws3.row_dimensions[1].height = 26

    hdrs3_general = ["CONDUCTOR", "TOTAL", "COMPLET.", "ANULADOS", "INCUMPL.", "EN CURSO", "% CUMPL."]
    fill_bloque1 = PatternFill("solid", start_color="203A43")
    for ci, h in enumerate(hdrs3_general, start=1):
        c = ws3.cell(2, ci, h)
        c.font = ft_header; c.fill = fill_bloque1
        c.alignment = centro; c.border = borde
    ws3.row_dimensions[2].height = 20

    ws3.column_dimensions["H"].width = 2
    hdrs3_tiempos = ["CONDUCTOR","ESPERA CARGUE\n(prom.)","TRÁNSITO\n(prom.)","DESCARGUE\n(prom.)","TOTAL OP.\n(prom.)"]
    fill_tiempos = PatternFill("solid", start_color="1A5276")
    col_inicio_tiempos = 9
    for ci, h in enumerate(hdrs3_tiempos, start=col_inicio_tiempos):
        c = ws3.cell(2, ci, h)
        c.font = ft_header; c.fill = fill_tiempos
        c.alignment = centro; c.border = borde
    ws3.row_dimensions[2].height = 30

    ws3.column_dimensions[get_column_letter(col_inicio_tiempos + 5)].width = 2
    col_inicio_lento = col_inicio_tiempos + 6
    hdrs3_lento = ["🐢 MÁS LENTO EN...","ESPERA CARGUE","TRÁNSITO","DESCARGUE","TOTAL OPERACIÓN"]
    fill_lento = PatternFill("solid", start_color="922B21")
    for ci, h in enumerate(hdrs3_lento, start=col_inicio_lento):
        c = ws3.cell(2, ci, h)
        c.font = ft_header; c.fill = fill_lento
        c.alignment = centro; c.border = borde

    ws3.column_dimensions[get_column_letter(col_inicio_lento + 5)].width = 2
    col_inicio_rapido = col_inicio_lento + 6
    hdrs3_rapido = ["⚡ MÁS RÁPIDO EN...","ESPERA CARGUE","TRÁNSITO","DESCARGUE","TOTAL OPERACIÓN"]
    fill_rapido = PatternFill("solid", start_color="1E8449")
    for ci, h in enumerate(hdrs3_rapido, start=col_inicio_rapido):
        c = ws3.cell(2, ci, h)
        c.font = ft_header; c.fill = fill_rapido
        c.alignment = centro; c.border = borde

    # Calcular tiempos por conductor usando días extra
    tiempos_por_conductor = {}
    if "conductor" in df.columns:
        for conductor_nombre, grupo in df.groupby("conductor"):
            if not conductor_nombre or str(conductor_nombre).strip() == "":
                continue
            esperas, transitos, descargues = [], [], []
            for _, r in grupo.iterrows():
                d_sc_ = int(r.get("dias_salida_cargue",  0) or 0)
                d_ld_ = int(r.get("dias_llegada_descargue", 0) or 0)
                d_sd_ = int(r.get("dias_salida_descargue", 0) or 0)
                e = calcular_duracion(r.get("hora_cita_cargue"),       r.get("hora_salida_cargue"),     d_sc_)
                t = calcular_duracion(r.get("hora_salida_cargue"),     r.get("hora_llegada_descargue"), d_ld_)
                d = calcular_duracion(r.get("hora_llegada_descargue"), r.get("hora_salida_descargue"),  d_sd_)
                if e is not None: esperas.append(e)
                if t is not None: transitos.append(t)
                if d is not None: descargues.append(d)

            prom_e = sum(esperas)    / len(esperas)    if esperas    else None
            prom_t = sum(transitos)  / len(transitos)  if transitos  else None
            prom_d = sum(descargues) / len(descargues) if descargues else None
            prom_tot = (prom_e + prom_t + prom_d) if (prom_e and prom_t and prom_d) else None
            tiempos_por_conductor[conductor_nombre] = {
                "espera": prom_e, "transito": prom_t, "descargue": prom_d, "total": prom_tot
            }

    if "conductor" in df.columns:
        df_cond = df.groupby("conductor").agg(
            total=("conductor","count"),
            comp=("estado", lambda x: x.str.contains("Completado", na=False).sum()),
            anul=("estado", lambda x: x.str.contains("Anulado",    na=False).sum()),
            incu=("estado", lambda x: x.str.contains("Incumplido", na=False).sum()),
            curs=("estado", lambda x: x.str.contains("En Curso",   na=False).sum()),
        ).reset_index().sort_values("total", ascending=False)

        for i, row in enumerate(df_cond.itertuples(), start=3):
            pct = f"{round(row.comp/row.total*100,1)}%" if row.total > 0 else "0%"
            vals = [row.conductor, row.total, row.comp, row.anul, row.incu, row.curs, pct]
            fill_c = PatternFill("solid", start_color="EBF5FB") if i % 2 == 0 else None
            for ci, v in enumerate(vals, start=1):
                c = ws3.cell(i, ci, v)
                c.font = ft_normal; c.border = borde
                c.alignment = izq if ci == 1 else centro
                if fill_c: c.fill = fill_c

    conductores_con_tiempos = sorted(tiempos_por_conductor.keys())
    for i, cond in enumerate(conductores_con_tiempos, start=3):
        t = tiempos_por_conductor[cond]
        fill_c = PatternFill("solid", start_color="D6EAF8") if i % 2 == 0 else None
        vals_t = [cond, mins_a_str(t["espera"]), mins_a_str(t["transito"]),
                  mins_a_str(t["descargue"]), mins_a_str(t["total"])]
        for ci, v in enumerate(vals_t, start=col_inicio_tiempos):
            c = ws3.cell(i, ci, v)
            c.font = ft_normal; c.border = borde
            c.alignment = izq if ci == col_inicio_tiempos else centro
            if fill_c: c.fill = fill_c

    etapas = ["espera", "transito", "descargue", "total"]

    def ranking_etapa(tiempos_dict, etapa, ascendente=False):
        datos = [(cond, d[etapa]) for cond, d in tiempos_dict.items() if d[etapa] is not None]
        datos.sort(key=lambda x: x[1], reverse=not ascendente)
        return datos

    for etapa_idx, etapa in enumerate(etapas):
        col_etapa_lento  = col_inicio_lento  + 1 + etapa_idx
        col_etapa_rapido = col_inicio_rapido + 1 + etapa_idx

        ranking_lento  = ranking_etapa(tiempos_por_conductor, etapa, ascendente=False)
        ranking_rapido = ranking_etapa(tiempos_por_conductor, etapa, ascendente=True)

        for fila_idx, (cond, mins) in enumerate(ranking_lento, start=3):
            c_nombre = ws3.cell(fila_idx, col_inicio_lento, cond)
            c_nombre.font = ft_normal; c_nombre.border = borde; c_nombre.alignment = izq
            fill_l = PatternFill("solid", start_color="FADBD8") if fila_idx % 2 == 0 else None
            if fill_l: c_nombre.fill = fill_l
            c_val = ws3.cell(fila_idx, col_etapa_lento, mins_a_str(mins))
            c_val.font = ft_normal; c_val.border = borde; c_val.alignment = centro
            if fila_idx == 3:
                c_val.font = Font(name="Calibri", bold=True, size=10, color="C0392B")
                c_nombre.font = Font(name="Calibri", bold=True, size=10, color="C0392B")
            elif fill_l:
                c_val.fill = fill_l

        for fila_idx, (cond, mins) in enumerate(ranking_rapido, start=3):
            c_nombre = ws3.cell(fila_idx, col_inicio_rapido, cond)
            c_nombre.font = ft_normal; c_nombre.border = borde; c_nombre.alignment = izq
            fill_r = PatternFill("solid", start_color="D5F5E3") if fila_idx % 2 == 0 else None
            if fill_r: c_nombre.fill = fill_r
            c_val = ws3.cell(fila_idx, col_etapa_rapido, mins_a_str(mins))
            c_val.font = ft_normal; c_val.border = borde; c_val.alignment = centro
            if fila_idx == 3:
                c_val.font = Font(name="Calibri", bold=True, size=10, color="1E8449")
                c_nombre.font = Font(name="Calibri", bold=True, size=10, color="1E8449")
            elif fill_r:
                c_val.fill = fill_r

    anchos_ws3 = {
        "A": 32, "B": 8, "C": 10, "D": 10, "E": 10, "F": 10, "G": 10,
        "H": 2, "I": 32, "J": 16, "K": 16, "L": 16, "M": 16,
        "N": 2, "O": 30, "P": 16, "Q": 16, "R": 16, "S": 16,
        "T": 2, "U": 30, "V": 16, "W": 16, "X": 16, "Y": 16,
    }
    for col_l, w in anchos_ws3.items():
        ws3.column_dimensions[col_l].width = w
    ws3.freeze_panes = "A3"

    # ==================== HOJA TIEMPOS ====================
    ws4 = wb.create_sheet("Tiempos")
    ws4["A1"] = "Analisis de Tiempos por Viaje"
    ws4["A1"].font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    ws4["A1"].fill = PatternFill("solid", start_color="0F2027")
    ws4["A1"].alignment = centro
    ws4.row_dimensions[1].height = 26

    hdrs4 = ["FECHA","PLACA","CONDUCTOR","CLIENTE","ESPERA CARGUE","TRANSITO","DESCARGUE","TOTAL OPERACION"]
    for ci, h in enumerate(hdrs4, start=1):
        c = ws4.cell(2, ci, h)
        c.font = ft_header
        c.fill = PatternFill("solid", start_color="203A43")
        c.alignment = centro; c.border = borde
    ws4.row_dimensions[2].height = 20

    tot_espera = tot_transito = tot_desc = tot_total = 0
    count_e = count_t = count_d = count_tot = 0

    for i, (_, row) in enumerate(df.iterrows(), start=3):
        d_sc_ = int(row.get("dias_salida_cargue",  0) or 0)
        d_ld_ = int(row.get("dias_llegada_descargue", 0) or 0)
        d_sd_ = int(row.get("dias_salida_descargue", 0) or 0)
        t_espera  = calcular_duracion(row.get("hora_cita_cargue"),       row.get("hora_salida_cargue"),     d_sc_)
        t_transit = calcular_duracion(row.get("hora_salida_cargue"),     row.get("hora_llegada_descargue"), d_ld_)
        t_desc    = calcular_duracion(row.get("hora_llegada_descargue"), row.get("hora_salida_descargue"),  d_sd_)
        t_total   = (t_espera + t_transit + t_desc) if (t_espera and t_transit and t_desc) else None

        if t_espera  is not None: tot_espera  += t_espera;  count_e   += 1
        if t_transit is not None: tot_transito += t_transit; count_t   += 1
        if t_desc    is not None: tot_desc    += t_desc;    count_d   += 1
        if t_total   is not None: tot_total   += t_total;   count_tot += 1

        vals = [
            str(row.get("fecha","")), str(row.get("placa","")),
            str(row.get("conductor","")), str(row.get("cliente","")),
            mins_a_str(t_espera), mins_a_str(t_transit),
            mins_a_str(t_desc), mins_a_str(t_total),
        ]
        fill_t = PatternFill("solid", start_color="EBF5FB") if i % 2 == 0 else None
        for ci, v in enumerate(vals, start=1):
            c = ws4.cell(i, ci, v)
            c.font = ft_normal; c.border = borde
            c.alignment = izq if ci in (1,2,3,4) else centro
            if fill_t: c.fill = fill_t

    fila_prom = len(df) + 3
    cp = ws4.cell(fila_prom, 1, "PROMEDIO")
    cp.font = ft_total; cp.fill = fill_total; cp.alignment = centro; cp.border = borde
    for ci, (tot, cnt) in enumerate([
        (tot_espera, count_e), (tot_transito, count_t),
        (tot_desc, count_d), (tot_total, count_tot)
    ], start=5):
        c = ws4.cell(fila_prom, ci, mins_a_str(tot/cnt if cnt > 0 else None))
        c.font = ft_total; c.fill = fill_total; c.alignment = centro; c.border = borde

    for col_l, w in zip(["A","B","C","D","E","F","G","H"], [12,10,28,20,16,16,16,18]):
        ws4.column_dimensions[col_l].width = w
    ws4.freeze_panes = "A3"

    # ==================== HOJA GRAFICA ESTADOS ====================
    try:
        from openpyxl.chart import PieChart, Reference
        from openpyxl.chart.series import DataPoint
        ws5 = wb.create_sheet("Grafica")
        ws5["A1"] = "Estado"; ws5["B1"] = "Cantidad"
        ws5["A1"].font = ft_header; ws5["B1"].font = ft_header
        ws5["A1"].fill = PatternFill("solid", start_color="203A43")
        ws5["B1"].fill = PatternFill("solid", start_color="203A43")
        estados_graf = ["Completado","Anulado","Incumplido","En Curso"]
        for i, est in enumerate(estados_graf, start=2):
            cnt = len(df[df["estado"].str.contains(est, na=False)]) if "estado" in df.columns else 0
            ws5.cell(i, 1, est).border = borde
            ws5.cell(i, 2, cnt).border  = borde
        pie = PieChart()
        pie.title = "Distribucion de Viajes por Estado"
        pie.style = 10
        labels = Reference(ws5, min_col=1, min_row=2, max_row=5)
        data   = Reference(ws5, min_col=2, min_row=1, max_row=5)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.width = 15; pie.height = 12
        colores = ["2ECC71","E74C3C","F39C12","3498DB"]
        for idx, color in enumerate(colores):
            pt = DataPoint(idx=idx)
            pt.graphicalProperties.solidFill = color
            pie.series[0].dPt.append(pt)
        ws5.add_chart(pie, "D1")
        for col_l, w in zip(["A","B"], [16, 10]):
            ws5.column_dimensions[col_l].width = w
    except Exception:
        pass

    ws.freeze_panes = "A3"
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ==================== MAIN ====================
def main():
    st.markdown("""
    <div class="main-header">
        <h1>🚚 CONTROL DE VIAJES</h1>
        <p>Registro y seguimiento de operaciones de transporte</p>
    </div>
    """, unsafe_allow_html=True)

    if "db" not in st.session_state:
        st.session_state.db = DB()
    if "editando_id" not in st.session_state:
        st.session_state.editando_id = None

    db = st.session_state.db

    tab1, tab2, tab3 = st.tabs(["📝 Nuevo Viaje", "🔍 Historial y Reportes", "📊 Dashboard"])

    # ===================== TAB 1: NUEVO VIAJE =====================
    with tab1:
        st.markdown("### Registrar Nuevo Viaje")

        f1, f2, f3, f4 = st.columns(4)
        with f1:
            fecha_pre = st.date_input("📅 Fecha", datetime.now(), key="pre_fecha")
        with f2:
            placas_lista = list(PLACA_CONDUCTOR.keys())
            placa_pre = st.selectbox("🚛 Placa", placas_lista, key="pre_placa")
        with f3:
            conductor_fijo = PLACA_CONDUCTOR.get(placa_pre)
            cond_opts = ["— Seleccionar —"] + TODOS_CONDUCTORES
            cond_default = cond_opts.index(conductor_fijo) if conductor_fijo in cond_opts else 0
            conductor_sel = st.selectbox("👤 Conductor", cond_opts, index=cond_default, key="pre_conductor")
        with f4:
            cli_sel = st.selectbox("🏢 Cliente", CLIENTES_FRECUENTES + [LABEL_MANUAL_CLI], key="pre_cliente")

        if cli_sel == LABEL_MANUAL_CLI:
            cliente_pre = st.text_input("✏️ Escribir cliente manualmente",
                                        placeholder="Nombre del cliente...", key="pre_cli_manual")
        else:
            cliente_pre = cli_sel

        st.markdown("#### 🗺️ Ruta")
        ruta_opts = [f"{o}  →  {d}" for o, d in RUTAS_FRECUENTES] + [LABEL_MANUAL]
        ruta_sel = st.selectbox("🗺️ Ruta frecuente", ruta_opts, index=len(ruta_opts)-1, key="pre_ruta")
        c5, c6 = st.columns(2)
        if ruta_sel == LABEL_MANUAL:
            with c5: origen_pre  = st.text_input("📍 Origen",  placeholder="Escribe el origen...",  key="pre_origen")
            with c6: destino_pre = st.text_input("🏁 Destino", placeholder="Escribe el destino...", key="pre_destino")
        else:
            _o, _d = ruta_sel.split("  →  ")
            with c5: st.info(f"📍 **Origen:** {_o}")
            with c6: st.info(f"🏁 **Destino:** {_d}")
            origen_pre, destino_pre = _o, _d

        # --- Horas fuera del form para preview en tiempo real ---
        hora_cita, hora_sc, hora_ld, hora_sd, d_sc, d_ld, d_sd = widget_horas(
            prefix="new",
            label_cita="Cita Cargue",
            label_sal_cargue="Salida Cargue",
            label_ll_desc="Llegada Descargue",
            label_sal_desc="Salida Descargue",
        )

        with st.form("form_viaje", clear_on_submit=True):
            fecha = fecha_pre
            placa = placa_pre
            conductor = "" if conductor_sel == "— Seleccionar —" else conductor_sel
            cliente = cliente_pre
            origen = origen_pre
            destino = destino_pre

            st.markdown("#### 📦 Información de Carga")
            d1, d2, d3, d4 = st.columns(4)
            with d1: contenedor         = st.text_input("Contenedor")
            with d2: carga              = st.text_input("Carga")
            with d3: numero_importacion = st.text_input("Nº Importación / BL")
            with d4: manifiesto         = st.text_input("Manifiesto")

            e1, e2 = st.columns([1, 3])
            with e1: estado      = st.selectbox("🚦 Estado", ESTADOS_VIAJE)
            with e2: observacion = st.text_area("📝 Observaciones", height=80)

            submitted = st.form_submit_button("💾 Guardar Viaje", type="primary", use_container_width=True)

        if submitted:
            if not placa:
                st.error("⚠️ La placa es obligatoria.")
            else:
                datos = {
                    "fecha": fecha, "placa": placa, "conductor": conductor,
                    "cliente": cliente, "origen": origen, "destino": destino,
                    "hora_cita_cargue": hora_cita,
                    "hora_salida_cargue": hora_sc,
                    "hora_llegada_descargue": hora_ld,
                    "hora_salida_descargue": hora_sd,
                    "contenedor": contenedor, "carga": carga,
                    "numero_importacion_bl": numero_importacion,
                    "manifiesto": manifiesto, "observacion": observacion,
                    "estado": estado.split(" ", 1)[1] if " " in estado else estado,
                    "dias_salida_cargue": d_sc,
                    "dias_llegada_descargue": d_ld,
                    "dias_salida_descargue": d_sd,
                }
                if db.guardar_viaje(datos):
                    st.success(f"✅ Viaje guardado — {placa} | {conductor} | {origen} → {destino}")
                    st.balloons()

    # ===================== TAB 2: HISTORIAL =====================
    with tab2:
        st.markdown("### 🔍 Historial de Viajes")

        with st.expander("🛠️ Filtros", expanded=True):
            f1, f2, f3, f4, f5, f6 = st.columns(6)
            with f1: fi   = st.date_input("Desde", datetime.now() - timedelta(days=30), key="h_fi")
            with f2: ff   = st.date_input("Hasta", datetime.now(), key="h_ff")
            with f3:
                placas_h = ["Todas"] + list(PLACA_CONDUCTOR.keys())
                fp = st.selectbox("Placa", placas_h, key="h_fp")
            with f4: fc   = st.text_input("Conductor", key="h_fc")
            with f5: fcli = st.text_input("Cliente",   key="h_fcli")
            with f6:
                estados_f = ["Todos"] + [e.split(" ", 1)[1] for e in ESTADOS_VIAJE]
                fe = st.selectbox("Estado", estados_f, key="h_fe")

        df = db.obtener_viajes(fi, ff, fp, fc, fcli, fe if fe != "Todos" else None)

        if not df.empty:
            k1, k2, k3, k4 = st.columns(4)
            k1.metric("Total Viajes", len(df))
            k2.metric("✅ Completados", len(df[df["estado"].str.contains("Completado", na=False)]))
            k3.metric("❌ Anulados",    len(df[df["estado"].str.contains("Anulado",    na=False)]))
            k4.metric("⚠️ Incumplidos", len(df[df["estado"].str.contains("Incumplido", na=False)]))

            st.divider()
            col_exp1, col_exp2 = st.columns([2, 5])
            with col_exp1:
                nombre_rep = st.text_input("Nombre del reporte", value="Control_Viajes", key="rep_nombre")
            with col_exp2:
                st.markdown("<br>", unsafe_allow_html=True)
                excel_data = generar_excel(df, titulo=nombre_rep)
                st.download_button(
                    "⬇️ Descargar Excel",
                    data=excel_data,
                    file_name=f"{nombre_rep}_{datetime.now(pytz.timezone('America/Bogota')).strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )

            st.divider()
            cols_tabla = ["id","fecha","placa","conductor","cliente","origen","destino",
                          "contenedor","carga","numero_importacion_bl","manifiesto","estado"]
            cols_ex = [c for c in cols_tabla if c in df.columns]
            st.dataframe(df[cols_ex], use_container_width=True, hide_index=True)

            st.divider()
            st.subheader("✏️ Ver Detalle / Editar")
            df["_label"] = df.apply(
                lambda r: f"ID {r['id']} | {r['fecha']} | {r['placa']} | {r.get('cliente','')} | {r.get('origen','')} → {r.get('destino','')} | {r.get('estado','')}",
                axis=1
            )
            sel = st.selectbox("Seleccionar viaje:", df["_label"].tolist(), key="h_sel")

            if sel:
                vid = int(sel.split(" | ")[0].replace("ID ", ""))
                row = df[df["id"] == vid].iloc[0]
                editando = st.session_state.editando_id == vid

                if not editando:
                    c1, c2, c3 = st.columns(3)
                    with c1:
                        st.info(f"**Placa:** {row['placa']}")
                        st.write(f"**Conductor:** {row.get('conductor','')}")
                        st.write(f"**Cliente:** {row.get('cliente','')}")
                        st.write(f"**Fecha:** {row['fecha']}")
                    with c2:
                        st.write(f"**Origen:** {row.get('origen','')}")
                        st.write(f"**Destino:** {row.get('destino','')}")
                        st.write(f"**Contenedor:** {row.get('contenedor','')}")
                        st.write(f"**Carga:** {row.get('carga','')}")
                    with c3:
                        st.write(f"**Importación/BL:** {row.get('numero_importacion_bl','')}")
                        st.write(f"**Manifiesto:** {row.get('manifiesto','')}")
                        estado_raw = str(row.get('estado',''))
                        color = "🟢" if "Completado" in estado_raw else ("🔴" if "Anulado" in estado_raw else "🟡")
                        st.write(f"**Estado:** {color} {estado_raw}")
                        st.write(f"**Observación:** {row.get('observacion','')}")

                    # Mostrar horas con días extra
                    d_sc_r = int(row.get("dias_salida_cargue",  0) or 0)
                    d_ld_r = int(row.get("dias_llegada_descargue", 0) or 0)
                    d_sd_r = int(row.get("dias_salida_descargue", 0) or 0)
                    sufsc = f" (+{d_sc_r}d)" if d_sc_r else ""
                    sufld = f" (+{d_ld_r}d)" if d_ld_r else ""
                    sufsd = f" (+{d_sd_r}d)" if d_sd_r else ""
                    st.write(
                        f"**Horas:** Cita: `{str_hora(row['hora_cita_cargue'])}` | "
                        f"Salida Cargue: `{str_hora(row['hora_salida_cargue'])}{sufsc}` | "
                        f"Llegada: `{str_hora(row['hora_llegada_descargue'])}{sufld}` | "
                        f"Salida Desc: `{str_hora(row['hora_salida_descargue'])}{sufsd}`"
                    )

                    # Mostrar tiempos calculados correctamente
                    t_e = calcular_duracion(row["hora_cita_cargue"],       row["hora_salida_cargue"],     d_sc_r)
                    t_t = calcular_duracion(row["hora_salida_cargue"],     row["hora_llegada_descargue"], d_ld_r)
                    t_d = calcular_duracion(row["hora_llegada_descargue"], row["hora_salida_descargue"],  d_sd_r)
                    t_tot = (t_e + t_t + t_d) if (t_e and t_t and t_d) else None
                    st.write(
                        f"**Tiempos:** ⏳ Espera: `{mins_a_str(t_e)}` | "
                        f"🚛 Tránsito: `{mins_a_str(t_t)}` | "
                        f"📦 Descargue: `{mins_a_str(t_d)}` | "
                        f"🕐 Total: `{mins_a_str(t_tot)}`"
                    )

                    bc1, bc2 = st.columns(2)
                    with bc1:
                        if st.button("✏️ Editar", key=f"eb_{vid}"):
                            st.session_state.editando_id = vid; st.rerun()
                    with bc2:
                        if st.button("🗑️ Eliminar", key=f"del_{vid}"):
                            db.eliminar_viaje(vid); st.success("Eliminado."); st.rerun()
                else:
                    st.markdown("#### ✏️ Editando viaje")

                    ec1, ec2, ec3, ec4 = st.columns(4)
                    with ec1: e_fecha = st.date_input("Fecha", value=row["fecha"], key=f"ef_{vid}")
                    with ec2:
                        placas_e = list(PLACA_CONDUCTOR.keys())
                        placa_idx = placas_e.index(row["placa"]) if row["placa"] in placas_e else 0
                        e_placa = st.selectbox("Placa", placas_e, index=placa_idx, key=f"ep_{vid}")
                    with ec3:
                        cond_fijo_e = PLACA_CONDUCTOR.get(e_placa)
                        cond_actual = str(row.get("conductor") or "")
                        cond_opts_e = ["— Seleccionar —"] + TODOS_CONDUCTORES
                        default_e = cond_opts_e.index(cond_fijo_e) if cond_fijo_e in cond_opts_e else (
                            cond_opts_e.index(cond_actual) if cond_actual in cond_opts_e else 0)
                        e_cond_sel = st.selectbox("👤 Conductor", cond_opts_e, index=default_e, key=f"ec_{vid}")
                        e_conductor = "" if e_cond_sel == "— Seleccionar —" else e_cond_sel
                    with ec4:
                        cli_actual = str(row.get("cliente") or "")
                        cli_opts = CLIENTES_FRECUENTES + [LABEL_MANUAL_CLI]
                        cli_idx = cli_opts.index(cli_actual) if cli_actual in cli_opts else len(cli_opts)-1
                        e_cli_sel = st.selectbox("Cliente", cli_opts, index=cli_idx, key=f"ecl_{vid}")
                        if e_cli_sel == LABEL_MANUAL_CLI:
                            e_cliente = st.text_input("Cliente (manual)",
                                                       value=cli_actual if cli_actual not in CLIENTES_FRECUENTES else "",
                                                       key=f"ecl_m_{vid}")
                        else:
                            e_cliente = e_cli_sel

                    er1, er2 = st.columns(2)
                    with er1: e_origen  = st.text_input("Origen",  value=str(row.get("origen")  or ""), key=f"eo_{vid}")
                    with er2: e_destino = st.text_input("Destino", value=str(row.get("destino") or ""), key=f"ed_{vid}")

                    # Widget de horas con días extra (pre-poblado con valores actuales)
                    e_hcc, e_hsc, e_hld, e_hsd, e_dsc, e_dld, e_dsd = widget_horas(
                        prefix=f"edit_{vid}",
                        label_cita="Cita Cargue",
                        label_sal_cargue="Salida Cargue",
                        label_ll_desc="Llegada Descargue",
                        label_sal_desc="Salida Descargue",
                        val_cita=hora_a_time(row["hora_cita_cargue"]),
                        val_sal_cargue=hora_a_time(row["hora_salida_cargue"]),
                        val_ll_desc=hora_a_time(row["hora_llegada_descargue"]),
                        val_sal_desc=hora_a_time(row["hora_salida_descargue"]),
                        dias_sc=int(row.get("dias_salida_cargue",  0) or 0),
                        dias_ld=int(row.get("dias_llegada_descargue", 0) or 0),
                        dias_sd=int(row.get("dias_salida_descargue", 0) or 0),
                    )

                    with st.form(f"edit_{vid}"):
                        ed1, ed2, ed3, ed4 = st.columns(4)
                        with ed1: e_cont  = st.text_input("Contenedor", value=str(row.get("contenedor") or ""),            key=f"eco_{vid}")
                        with ed2: e_carga = st.text_input("Carga",      value=str(row.get("carga") or ""),                 key=f"eca_{vid}")
                        with ed3: e_bl    = st.text_input("Imp / BL",   value=str(row.get("numero_importacion_bl") or ""), key=f"ebl_{vid}")
                        with ed4: e_man   = st.text_input("Manifiesto", value=str(row.get("manifiesto") or ""),            key=f"ema_{vid}")

                        estados_l = [e.split(" ", 1)[1] for e in ESTADOS_VIAJE]
                        est_actual = str(row.get("estado") or "Completado")
                        est_idx = estados_l.index(est_actual) if est_actual in estados_l else 0
                        ee1, ee2 = st.columns([1, 3])
                        with ee1: e_estado = st.selectbox("Estado", ESTADOS_VIAJE, index=est_idx, key=f"est_{vid}")
                        with ee2: e_obs    = st.text_area("Observaciones", value=str(row.get("observacion") or ""),
                                                           key=f"eob_{vid}", height=80)

                        sg1, sg2 = st.columns(2)
                        with sg1: guardar  = st.form_submit_button("💾 Guardar Cambios", type="primary")
                        with sg2: cancelar = st.form_submit_button("❌ Cancelar")

                    if guardar:
                        datos_edit = {
                            "fecha": e_fecha, "placa": e_placa, "conductor": e_conductor,
                            "cliente": e_cliente, "origen": e_origen, "destino": e_destino,
                            "hora_cita_cargue": e_hcc, "hora_salida_cargue": e_hsc,
                            "hora_llegada_descargue": e_hld, "hora_salida_descargue": e_hsd,
                            "contenedor": e_cont, "carga": e_carga,
                            "numero_importacion_bl": e_bl, "manifiesto": e_man,
                            "observacion": e_obs,
                            "estado": e_estado.split(" ", 1)[1] if " " in e_estado else e_estado,
                            "dias_salida_cargue": e_dsc,
                            "dias_llegada_descargue": e_dld,
                            "dias_salida_descargue": e_dsd,
                        }
                        if db.actualizar_viaje(vid, datos_edit):
                            st.success("✅ Viaje actualizado.")
                            st.session_state.editando_id = None; st.rerun()
                    if cancelar:
                        st.session_state.editando_id = None; st.rerun()
        else:
            st.warning("No hay viajes con los filtros seleccionados.")

    # ===================== TAB 3: DASHBOARD =====================
    with tab3:
        st.markdown("### 📊 Dashboard de Operaciones")

        try:
            import plotly.express as px
            import plotly.graph_objects as go

            col_r1, col_r2 = st.columns([2, 4])
            with col_r1:
                rango = st.date_input(
                    "Período",
                    value=(datetime.now().replace(day=1), datetime.now()),
                    key="dash_rango"
                )

            if not (isinstance(rango, (list, tuple)) and len(rango) == 2):
                st.info("Selecciona un rango de fechas completo.")
                return

            df_s = db.stats_dashboard(rango[0], rango[1])

            if df_s.empty:
                st.info("No hay datos en este período.")
                return

            total = len(df_s)
            comp  = len(df_s[df_s["estado"].str.contains("Completado", na=False)])
            anul  = len(df_s[df_s["estado"].str.contains("Anulado",    na=False)])
            incum = len(df_s[df_s["estado"].str.contains("Incumplido", na=False)])
            curso = len(df_s[df_s["estado"].str.contains("En Curso",   na=False)])
            pct   = round(comp / total * 100) if total > 0 else 0

            k1, k2, k3, k4, k5 = st.columns(5)
            k1.metric("🚚 Total Viajes", total)
            k2.metric("✅ Completados", comp, f"{pct}%")
            k3.metric("❌ Anulados", anul)
            k4.metric("⚠️ Incumplidos", incum)
            k5.metric("🔄 En Curso", curso)

            st.divider()

            g1, g2 = st.columns(2)
            with g1:
                st.markdown("#### Distribución por Estado")
                est_c = df_s["estado"].value_counts().reset_index()
                est_c.columns = ["estado", "cantidad"]
                colores_estado = {
                    "Completado": "#2ecc71", "Anulado": "#e74c3c",
                    "Incumplido": "#f39c12", "En Curso": "#3498db"
                }
                fig1 = px.pie(est_c, values="cantidad", names="estado", hole=0.45,
                              color="estado", color_discrete_map=colores_estado)
                fig1.update_layout(margin=dict(t=10, b=10), height=300)
                st.plotly_chart(fig1, use_container_width=True)

            with g2:
                st.markdown("#### Viajes por Día")
                df_dia = df_s.groupby("fecha").size().reset_index(name="viajes")
                fig2 = px.bar(df_dia, x="fecha", y="viajes",
                              color_discrete_sequence=["#2c5364"], text="viajes")
                fig2.update_traces(textposition="outside")
                fig2.update_layout(margin=dict(t=10, b=10), height=300,
                                   xaxis_title="", yaxis_title="Viajes")
                st.plotly_chart(fig2, use_container_width=True)

            st.divider()

            g3, g4 = st.columns(2)
            with g3:
                st.markdown("#### Viajes por Cliente")
                if "cliente" in df_s.columns and df_s["cliente"].notna().any():
                    df_cli = df_s.groupby("cliente").size().reset_index(name="viajes").sort_values("viajes")
                    fig3 = px.bar(df_cli, x="viajes", y="cliente", orientation="h",
                                  color="viajes", color_continuous_scale="Blues", text="viajes")
                    fig3.update_traces(textposition="outside")
                    fig3.update_layout(margin=dict(t=10, b=10), height=max(250, len(df_cli)*40),
                                       coloraxis_showscale=False, yaxis_title="", xaxis_title="Viajes")
                    st.plotly_chart(fig3, use_container_width=True)
                else:
                    st.info("Sin datos de cliente.")

            with g4:
                st.markdown("#### Viajes por Placa")
                df_placa = df_s.groupby("placa").size().reset_index(name="viajes").sort_values("viajes")
                fig4 = px.bar(df_placa, x="viajes", y="placa", orientation="h",
                              color="viajes", color_continuous_scale="Teal", text="viajes")
                fig4.update_traces(textposition="outside")
                fig4.update_layout(margin=dict(t=10, b=10), height=max(250, len(df_placa)*40),
                                   coloraxis_showscale=False, yaxis_title="", xaxis_title="Viajes")
                st.plotly_chart(fig4, use_container_width=True)

            st.divider()

            g5, g6 = st.columns(2)
            with g5:
                st.markdown("#### ⏱️ Tiempos Promedio de Operación")
                tiempos = []
                for _, r in df_s.iterrows():
                    d_sc_ = int(r.get("dias_salida_cargue",  0) or 0)
                    d_ld_ = int(r.get("dias_llegada_descargue", 0) or 0)
                    d_sd_ = int(r.get("dias_salida_descargue", 0) or 0)
                    t_cargue   = calcular_duracion(r["hora_cita_cargue"],       r["hora_salida_cargue"],     d_sc_)
                    t_transito = calcular_duracion(r["hora_salida_cargue"],     r["hora_llegada_descargue"], d_ld_)
                    t_descargue= calcular_duracion(r["hora_llegada_descargue"], r["hora_salida_descargue"],  d_sd_)
                    tiempos.append({
                        "espera_cargue": t_cargue,
                        "transito":      t_transito,
                        "descargue":     t_descargue
                    })
                df_t = pd.DataFrame(tiempos)
                prom = {
                    "Espera en Cargue": df_t["espera_cargue"].dropna().mean(),
                    "Tránsito":         df_t["transito"].dropna().mean(),
                    "Descargue":        df_t["descargue"].dropna().mean(),
                }
                prom_df = pd.DataFrame([
                    {"Etapa": k, "Minutos": round(v) if not pd.isna(v) else 0, "Tiempo": mins_a_str(v)}
                    for k, v in prom.items()
                ])
                fig5 = px.bar(prom_df, x="Etapa", y="Minutos",
                              color="Etapa", text="Tiempo",
                              color_discrete_sequence=["#2c5364", "#2980b9", "#1abc9c"])
                fig5.update_traces(textposition="outside")
                fig5.update_layout(margin=dict(t=10, b=10), height=300,
                                   showlegend=False, xaxis_title="", yaxis_title="Minutos promedio")
                st.plotly_chart(fig5, use_container_width=True)

            with g6:
                st.markdown("#### 📅 Ranking por Día de la Semana")
                df_s["dia_semana"] = pd.to_datetime(df_s["fecha"]).dt.day_name()
                orden = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"]
                nombres_es = {
                    "Monday":"Lunes","Tuesday":"Martes","Wednesday":"Miércoles",
                    "Thursday":"Jueves","Friday":"Viernes","Saturday":"Sábado","Sunday":"Domingo"
                }
                df_semana = df_s.groupby("dia_semana").size().reset_index(name="viajes")
                df_semana["orden"] = df_semana["dia_semana"].map({d: i for i, d in enumerate(orden)})
                df_semana = df_semana.sort_values("orden")
                df_semana["dia_es"] = df_semana["dia_semana"].map(nombres_es)
                fig6 = px.bar(df_semana, x="dia_es", y="viajes",
                              color="viajes", color_continuous_scale="Oranges", text="viajes")
                fig6.update_traces(textposition="outside")
                fig6.update_layout(margin=dict(t=10, b=10), height=300,
                                   coloraxis_showscale=False, xaxis_title="", yaxis_title="Viajes")
                st.plotly_chart(fig6, use_container_width=True)

            st.divider()

            st.markdown("#### 🏆 Ranking de Conductores")
            df_cond = df_s[
                df_s["conductor"].notna() & (df_s["conductor"].str.strip() != "")
            ].groupby("conductor").agg(
                viajes=("conductor", "count"),
                completados=("estado", lambda x: x.str.contains("Completado", na=False).sum()),
                anulados=("estado", lambda x: x.str.contains("Anulado", na=False).sum()),
                incumplidos=("estado", lambda x: x.str.contains("Incumplido", na=False).sum()),
            ).reset_index().sort_values("viajes", ascending=False).drop_duplicates(subset="conductor")
            df_cond["% Cumplimiento"] = (df_cond["completados"] / df_cond["viajes"] * 100).round(1).astype(str) + "%"
            df_cond.columns = ["Conductor", "Total", "✅ Comp.", "❌ Anul.", "⚠️ Incump.", "% Cumplimiento"]
            st.dataframe(df_cond, use_container_width=True, hide_index=True)

        except ImportError:
            st.warning("Instala plotly: `pip install plotly`")
        except Exception as e:
            st.error(f"Error en dashboard: {e}")


if __name__ == "__main__":
    main()
